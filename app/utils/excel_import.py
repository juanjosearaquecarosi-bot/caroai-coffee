"""
excel_import.py  —  Importa productos desde el Excel 'Caroai Cafe control ventas'.

Uso:  python -m app.utils.excel_import

Lee la hoja 'Precios' del Excel y extrae:
- Bebidas: nombre + precio COP + USD + Bs (columnas A-D)
- Comidas: nombre + precio COP + USD + Bs (columnas E-H)

Luego inserta o actualiza los productos en la base de datos.
"""

import os
import openpyxl
from app.models import db, Producto
from app import create_app

EXCEL_FILENAME = 'Caroai Cafe control ventas (1).xlsx'


def find_excel():
    start = os.path.dirname(os.path.abspath(__file__))
    for _ in range(5):
        candidate = os.path.join(start, EXCEL_FILENAME)
        if os.path.exists(candidate):
            return candidate
        start = os.path.dirname(start)
    return None


def parse_excel(excel_path):
    """
    Lee la hoja 'Precios' y extrae productos con nombre, tipo y precios.

    Estructura:
      Fila 1: Tasa BCV
      Fila 2: Cambio TRM
      Fila 3: Bebidas | Pesos | Dolares | Precio Bs | Comida | Pesos | Dolares | Precio Bs
      Filas 4+: datos

    Columnas bebidas: A=nombre, B=COP, C=USD, D=Bs
    Columnas comidas: E=nombre, F=COP, G=USD, H=Bs
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    productos = []
    sheet_name = 'Precios' if 'Precios' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Leer bebidas (cols A-D)
    for row_idx in range(4, ws.max_row + 1):
        nombre = ws.cell(row=row_idx, column=1).value
        precio_cop = ws.cell(row=row_idx, column=2).value
        precio_usd = ws.cell(row=row_idx, column=3).value
        precio_bs = ws.cell(row=row_idx, column=4).value

        nombre_str = str(nombre).strip() if nombre else ''

        # Detectar fin de sección o saltar encabezados
        if not nombre_str or len(nombre_str) < 2:
            # Podría ser comienzo de comidas, salir
            continue
        if any(kw in nombre_str.lower() for kw in ['tasa', 'cambio', 'dolar', 'precio', 'bebidas', 'comida', 'producto']):
            continue

        cop_val = _parse_precio(precio_cop)
        usd_val = _parse_precio_float(precio_usd)
        bs_val = _parse_precio_float(precio_bs)

        if cop_val and cop_val > 0:
            tipo = 'cerveza' if any(kw in nombre_str.lower() for kw in ['cerveza', 'malta']) else 'grano' if any(kw in nombre_str.lower() for kw in ['kilo', 'gramo', 'kg', 'origen']) else 'bebida'
            productos.append({
                'nombre': nombre_str,
                'tipo': tipo,
                'precio_cop': cop_val,
                'precio_usd': usd_val if usd_val and usd_val > 0 else None,
                'precio_bs': bs_val if bs_val and bs_val > 0 else None,
            })

    # Leer comidas (cols E-H)
    for row_idx in range(4, ws.max_row + 1):
        nombre = ws.cell(row=row_idx, column=5).value
        precio_cop = ws.cell(row=row_idx, column=6).value
        precio_usd = ws.cell(row=row_idx, column=7).value
        precio_bs = ws.cell(row=row_idx, column=8).value

        nombre_str = str(nombre).strip() if nombre else ''
        if not nombre_str or len(nombre_str) < 2:
            continue
        if any(kw in nombre_str.lower() for kw in ['tasa', 'cambio', 'dolar', 'comida', 'producto']):
            continue

        cop_val = _parse_precio(precio_cop)
        usd_val = _parse_precio_float(precio_usd)
        bs_val = _parse_precio_float(precio_bs)

        if cop_val and cop_val > 0:
            productos.append({
                'nombre': nombre_str,
                'tipo': 'comida',
                'precio_cop': cop_val,
                'precio_usd': usd_val if usd_val and usd_val > 0 else None,
                'precio_bs': bs_val if bs_val and bs_val > 0 else None,
            })

    return productos


def _parse_precio(val):
    if val is None:
        return 0
    try:
        # Handle strings like "4,000" or "$4,000"
        s = str(val).replace(',', '').replace('$', '').replace(' ', '').strip()
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _parse_precio_float(val):
    if val is None:
        return None
    try:
        s = str(val).replace(',', '').replace('$', '').replace(' ', '').strip()
        if s in ('0', '0.0', '0.00', ''):
            return None
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def import_productos(excel_path=None):
    if excel_path is None:
        excel_path = find_excel()

    if not excel_path:
        print(f'❌ Excel "{EXCEL_FILENAME}" no encontrado.')
        return 0, 0

    print(f'📄 Leyendo: {excel_path}')
    productos_data = parse_excel(excel_path)
    if not productos_data:
        print('   ⚠️  No se encontraron productos con precio > 0 en el Excel.')
        return 0, 0
    print(f'   → {len(productos_data)} productos con precio')

    insertados = 0
    actualizados = 0

    for data in productos_data:
        nombre = data['nombre']
        existente = Producto.query.filter_by(nombre=nombre).first()
        if existente:
            changed = False
            if existente.precio_cop != data['precio_cop']:
                existente.precio_cop = data['precio_cop']
                existente.precio_venta_cop = data['precio_cop']
                changed = True
            if data.get('precio_usd') is not None and existente.precio_usd != data['precio_usd']:
                existente.precio_usd = data['precio_usd']
                changed = True
            if data.get('precio_bs') is not None and existente.precio_bs != data['precio_bs']:
                existente.precio_bs = data['precio_bs']
                changed = True
            if changed:
                actualizados += 1
        else:
            prod = Producto(
                nombre=nombre,
                tipo=data['tipo'],
                precio_cop=data['precio_cop'],
                precio_venta_cop=data['precio_cop'],
                precio_usd=data.get('precio_usd'),
                precio_bs=data.get('precio_bs'),
                descuenta_inventario=False,
            )
            db.session.add(prod)
            insertados += 1

    db.session.commit()

    total = Producto.query.count()
    print(f'\n📊 Resumen:')
    print(f'   Insertados: {insertados}')
    print(f'   Actualizados: {actualizados}')
    print(f'   Total en BD: {total}')

    return insertados, actualizados


if __name__ == '__main__':
    app = create_app()
    with app.app_context():
        import_productos()
